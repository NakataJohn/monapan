#github
import os,openpyxl,pprint
from copy import copy
from openpyxl.utils import column_index_from_string,get_column_letter
#os.chdir('E:\文档')
print("Opening workbook................")
wb=openpyxl.load_workbook('nosol.xlsx')
sheet1=wb.get_sheet_by_name('Sheet1')
areacodes=[11,13,18,19,10,91,90,97,31,34,36,30,38,75,17,76,71,74,51,59,50,83,81,85,86,79,84,87,70,88,89]
tags=['1-10','11-30','31-50','51-100','101-300','300+','nohand']

nosol={}

for a in areacodes:
	nosol.setdefault(str(a),{})
	for t in tags:
		nosol[str(a)].setdefault(t,0)
print("Please ingore something like this,a new dir is creating")
		
print("Reading rows....................")
for rows in range(1,sheet1.max_row+1):
	areacode = str(sheet1['A'+str(rows)].value)
	tag = sheet1['B'+str(rows)].value
	count = int(sheet1['C'+str(rows)].value)
	nosol[areacode][tag]=count

pprint.pprint(nosol)

#写入py文件
#print("Writing results....")
#resultfile = open('renosol.py','w')
#resultfile.write('data='+pprint.pformat(nosol))
#resultfile.close()
#print("Done!")	

#写入excel文件
nwb=openpyxl.Workbook()
sheet=nwb.get_active_sheet()
a1=copy(areacodes)
t1=copy(tags)
#写表头
for r in range(len(a1)):
	sheet['A'+str(r+2)]=str(a1[r])

for c in range(2,9): 
	sheet[get_column_letter(c)+'1']=t1[c-2]

#插入数据
for r in range(2,sheet.max_row+1):
	for i in range(2,sheet.max_column+1):
		sheet[get_column_letter(i)+str(r)]=nosol[sheet['A'+str(r)].value][sheet[get_column_letter(i)+'1'].value]
	
nwb.save('result.xlsx')

#修复了一些bug

	
	
	

